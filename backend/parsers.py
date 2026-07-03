"""
Parseurs d'imports bancaires -> schéma unifié.

Schéma cible d'une opération :
    date (datetime), libelle (str), montant (float signé : - dépense / + revenu),
    categorie (str), sous_categorie (str), source (str), compte (str), op_id (str)

Le montant est TOUJOURS signé : négatif = sortie d'argent, positif = entrée.
op_id est une empreinte stable utilisée pour le dédoublonnage lors de ré-imports.
"""
from __future__ import annotations
import re
import csv
import hashlib
import datetime as dt
from io import BytesIO, StringIO

import pandas as pd
from openpyxl import load_workbook


def _norm(s) -> str:
    """Nettoie un libellé : espaces multiples / sauts de ligne -> un seul espace."""
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s)).strip()


def make_op_id(date: dt.datetime, libelle: str, montant: float, source: str,
               compte: str = "", seq: int = 0) -> str:
    """Empreinte stable d'une opération pour éviter les doublons au ré-import.
    Inclut le compte (la même opération peut exister sur deux comptes) et un
    n° d'occurrence `seq` (deux achats identiques le même jour sont distincts)."""
    base = (f"{source}|{date.date().isoformat()}|{_norm(libelle).lower()}"
            f"|{montant:.2f}|{_norm(compte).lower()}|{seq}")
    return hashlib.sha1(base.encode("utf-8")).hexdigest()[:16]


def _seq_counter():
    """Compteur d'occurrences (date, libellé, montant, compte) au sein d'un fichier."""
    seen: dict = {}
    def next_seq(date: dt.datetime, lib: str, montant: float, compte: str) -> int:
        key = (date.date(), _norm(lib).lower(), round(montant, 2), _norm(compte).lower())
        seen[key] = seen.get(key, -1) + 1
        return seen[key]
    return next_seq


# --------------------------------------------------------------------------- #
# BoursoBank / Boursorama : CSV séparé par ';', décimales à la virgule,
# montant déjà signé, et déjà catégorisé (categoryParent / category).
# --------------------------------------------------------------------------- #
def parse_boursorama(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig")
    df = pd.read_csv(StringIO(text), sep=";", dtype=str).fillna("")
    df["montant"] = (
        df["amount"].str.replace(".", "", regex=False)
        .str.replace(",", ".", regex=False)
        .replace("", "0")
        .astype(float)
    )
    ops, next_seq = [], _seq_counter()
    for _, r in df.iterrows():
        date = pd.to_datetime(r["dateOp"]).to_pydatetime()
        lib = _norm(r["label"])
        montant = float(r["montant"])
        cat = _norm(r["categoryParent"]) or "Non catégorisé"
        compte = _norm(r["accountLabel"]) or "BoursoBank"
        ops.append({
            "op_id": make_op_id(date, lib, montant, "BoursoBank", compte,
                                next_seq(date, lib, montant, compte)),
            "date": date,
            "libelle": lib,
            "montant": round(montant, 2),
            "categorie": cat,
            "sous_categorie": _norm(r["category"]),
            "source": "BoursoBank",
            "compte": compte,
            "categorized_by": "bank",
        })
    return ops


# --------------------------------------------------------------------------- #
# Crédit Agricole : XLSX, colonnes Débit / Crédit séparées, décimales au point,
# en-tête déclarant le type de compte + titulaire + solde, plusieurs blocs.
# Une ligne n'est une opération que si sa 1re cellule est une vraie date.
# --------------------------------------------------------------------------- #
def _ca_type(label: str) -> tuple[str, str]:
    """Déduit (type_actif, nom_affiché) depuis le libellé de compte de l'en-tête."""
    l = label.lower()
    nom = _norm(label.split("n°")[0]) or label
    if "livret jeune" in l:            return "livret_jeune", "Livret Jeune"
    if "livret a" in l:                return "livret_a", "Livret A"
    if "livret" in l:                  return "autre", nom
    if "pea" in l:                     return "pea", "PEA"
    if "compte" in l or "dépôt" in l:  return "compte_courant", nom
    return "autre", nom


def _amount(cell: str):
    """Parse un montant type ' 1 615,60 €' / '16\\u202f721,82 €' -> float.
    Supprime tout sauf chiffres/virgule/point/signe : robuste aux espaces
    insécables (U+00A0) et fines insécables (U+202F) que la banque utilise
    comme séparateurs de milliers."""
    v = re.sub(r"[^\d,.\-]", "", str(cell))
    if "," in v:   # format français : le point éventuel est un séparateur de milliers
        v = v.replace(".", "").replace(",", ".")
    try:
        return round(float(v), 2)
    except ValueError:
        return None


def _is_account_header(cell: str) -> bool:
    cl = cell.lower()
    return "n°" in cl and ("compte" in cl or "livret" in cl or "pea" in cl or "dépôt" in cl)


def _ca_parse_all(content: bytes) -> dict:
    """Parse un export CA pouvant contenir PLUSIEURS comptes (chaque bloc a son
    en-tête '<Type> n° …', son 'Solde au …' puis ses opérations).
    -> {ops:[...], accounts:[{type,nom,solde}], holder_tokens:[...]}"""
    wb = load_workbook(BytesIO(content), read_only=True)
    rows = list(wb.active.iter_rows(values_only=True))

    # Titulaire (1re ligne M./MME…) -> nom+prénom pour repérer les virements internes.
    holder = ""
    for r in rows[:15]:
        for c in r:
            if c and str(c).strip().startswith(("M.", "MME", "MLE", "M ")):
                holder = _norm(c); break
        if holder:
            break
    sans_civ = re.sub(r"^(m\.|mme|mle|m)\s+", "", holder.lower()).split()
    holder_tokens = [t for t in sans_civ[:2] if len(t) >= 3]

    ops, accounts = [], []
    next_seq = _seq_counter()
    cur_type, cur_nom = "compte_courant", "Compte de Dépôt"
    cur_acc = None
    for r in rows:
        if not r:
            continue
        # Ligne d'opération : 1re cellule = vraie date -> rattachée au compte courant.
        if isinstance(r[0], dt.datetime):
            date, lib, debit, credit = (list(r) + [None] * 4)[:4]
            deb = float(debit) if debit not in (None, "") else 0.0
            cre = float(credit) if credit not in (None, "") else 0.0
            montant = round(cre - deb, 2)
            lib = _norm(lib)
            low = lib.lower()
            is_transfer = ("virement" in low and holder_tokens
                           and all(tok in low for tok in holder_tokens))
            ops.append({
                "op_id": make_op_id(date, lib, montant, "Crédit Agricole", cur_nom,
                                    next_seq(date, lib, montant, cur_nom)),
                "date": date, "libelle": lib, "montant": montant,
                "categorie": "Virements internes" if is_transfer else "Non catégorisé",
                "sous_categorie": "", "source": "Crédit Agricole", "compte": cur_nom,
                "categorized_by": "rule" if is_transfer else "none",
            })
            continue
        # Ligne de métadonnées : en-tête de compte ? solde ?
        cells = [str(c) for c in r if c is not None]
        for c in cells:
            if _is_account_header(c):
                cur_type, cur_nom = _ca_type(c)
                cur_acc = {"type": cur_type, "nom": cur_nom, "solde": None}
                accounts.append(cur_acc)
                break
        for i, c in enumerate(cells):
            if "solde au" in c.lower() and i + 1 < len(cells):
                val = _amount(cells[i + 1])
                if val is not None:
                    if cur_acc is None:   # solde avant tout en-tête -> compte par défaut
                        cur_acc = {"type": cur_type, "nom": cur_nom, "solde": None}
                        accounts.append(cur_acc)
                    cur_acc["solde"] = val

    return {"ops": ops, "accounts": accounts, "holder_tokens": holder_tokens}


def parse_credit_agricole(content: bytes) -> list[dict]:
    return _ca_parse_all(content)["ops"]


# --------------------------------------------------------------------------- #
# Soldes : renvoie TOUS les comptes présents (compte courant + livrets + PEA).
# --------------------------------------------------------------------------- #
def extract_balances(filename: str, content: bytes) -> list[dict]:
    """-> [{compte, source, solde, type}] pour chaque compte de l'export."""
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xls")) or content[:2] == b"PK":
        accounts = _ca_parse_all(content)["accounts"]
        return [{"compte": a["nom"], "source": "Crédit Agricole",
                 "solde": a["solde"], "type": a["type"]}
                for a in accounts if a["solde"] is not None]
    # BoursoBank CSV : dernier solde par compte (toujours un compte courant)
    text = content.decode("utf-8-sig")
    df = pd.read_csv(StringIO(text), sep=";", dtype=str).fillna("")
    df["dateOp"] = pd.to_datetime(df["dateOp"])
    out = []
    for compte, sub in df.sort_values("dateOp", ascending=False).groupby("accountLabel"):
        bal = _amount(sub.iloc[0]["accountbalance"])
        if bal is not None:
            out.append({"compte": _norm(compte) or "BoursoBank",
                        "source": "BoursoBank", "solde": bal,
                        "type": "compte_courant"})
    return out


def detect_and_parse(filename: str, content: bytes) -> list[dict]:
    name = (filename or "").lower()
    # XLSX = archive ZIP -> commence par 'PK'
    if name.endswith((".xlsx", ".xls")) or content[:2] == b"PK":
        return parse_credit_agricole(content)
    # Sinon CSV : on regarde l'entête
    head = content[:300].decode("utf-8-sig", errors="ignore").lower()
    if "categoryparent" in head or "dateop" in head:
        return parse_boursorama(content)
    raise ValueError(
        "Format non reconnu. Attendu : export CSV BoursoBank ou XLSX Crédit Agricole."
    )
